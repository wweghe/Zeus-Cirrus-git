from typing import Any, Dict, List, Union

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import load_workbook
from datetime import datetime
import shutil, os
from pathlib import Path

import common.constants as constants
import common.utils as utils

from domain.batch_run_result import BatchRunResult
from domain.batch_config import BatchConfig


class BatchRunReportService:

    # private members
    def __init__(self) -> None:
        pass

    
    def __get_last_column_index(self, 
                                sheet : Worksheet,
                                start_column_index : int = 0) -> int:
        result = start_column_index
        for col in sheet.iter_cols(min_col = start_column_index, max_row = 1):
            headerCell = col[0]

            if (headerCell.value is not None):
                result = headerCell.col_idx + 1
            else:
                break
        
        return result
    

    def __get_column_map(self, 
                         sheet : Worksheet,
                         column_list : List[str]
                        ) -> Dict[str, int]:
        column_map : Dict[str, int] = {}

        for col in sheet.iter_cols(max_row=1):
            headerCell = col[0]

            if (headerCell.value in column_list):
                col_idx = headerCell.col_idx - 1
                column_map.update({ headerCell.value: col_idx })
        
        return column_map
    
    
    def __get_row_for_object(self, 
                             sheet : Worksheet,
                             column_map : Dict[str, int],
                             key_column_map : Dict[str, Any],
                             row_offset : int = 2
                            ) -> int:
        for row in sheet.iter_rows(min_row = row_offset):
            row_idx = row[0].row
            match = []
            for col in column_map.keys():
                cell_value = row[column_map[col]].value
                match += [cell_value == key_column_map[col]]
            if (False not in match):
                return row_idx
        
        return -1
    

    def _write_status(self, 
                      batch_run_results : List[BatchRunResult],
                      sheet : Worksheet):
        column : Union[int, None] = None
        column_map : Dict[str, int] = None
        row_offset = 2

        for result in batch_run_results:
            key_column_map = { \
                "objectId": result.object_id, 
                "sourceSystemCd": result.source_system_cd, 
                "action": result.action
                }
            if column_map is None:
                column_map = self.__get_column_map(sheet, column_list = list(key_column_map.keys()))
                column = self.__get_last_column_index(
                    sheet = sheet, 
                    start_column_index = max(column_map.values()))
                sheet.cell(row = 1, column = column).value = "SUCCESS"
                sheet.cell(row = 1, column = column + 1).value = "SKIP"
                sheet.cell(row = 1, column = column + 2).value = "ERROR"
                sheet.cell(row = 1, column = column + 3).value = "DURATION"

            row = self.__get_row_for_object(sheet, column_map, key_column_map = key_column_map, row_offset = row_offset)
            if (row > 0):
                
                sheet.cell(row = row, column = column).value = result.is_success if not result.is_skip else None
                sheet.cell(row = row, column = column + 1).value = result.is_skip
                sheet.cell(row = row, column = column + 2).value = result.error_message
                sheet.cell(row = row, column = column + 3).value = result.elapsed_time_str

                row_offset = row + 1
    

    # public members
    def create_log_report(self, 
                          batch_config : BatchConfig,
                          batch_run_results : List[BatchRunResult],
                          report_dir_path : str = None,
                          report_file_name : str = None
                          ) -> str:
        if batch_config is None: ValueError(f"batch_config cannot be empty")
        report_dir_path = utils.get_dir_path(report_dir_path)

        if report_file_name is None:
            path = Path(batch_config.file_path)
            report_file_path = f"{report_dir_path}/{path.stem}_results_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
        else:
            report_file_path = f"{report_dir_path}/{report_file_name}"

        shutil.copy(src = batch_config.file_path, 
                    dst = report_file_path)
        wb = load_workbook(report_file_path)
        # sheets_to_delete = [name for name in wb.sheetnames \
        #                     if name not in constants.SHEET_NAME_OBJECT_TYPE_RUNNABLE_MAP.keys()]
        # remove unnecesary sheets
        # for sheet_name in sheets_to_delete:
        #     del wb[sheet_name]
        # write status for each supported object type
        sheets = [sheetname.lower() for sheetname in wb.sheetnames]
        for sheet_name in constants.SHEET_NAME_OBJECT_TYPE_RUNNABLE_MAP.keys():
            if sheet_name in sheets:
                self._write_status(
                    batch_run_results = [res for res in batch_run_results if res.object_type == constants.SHEET_NAME_OBJECT_TYPE_RUNNABLE_MAP[sheet_name]],
                    sheet = wb[sheet_name])
        
        wb.save(filename = report_file_path)
        
        return report_file_path


            
            

        
        

