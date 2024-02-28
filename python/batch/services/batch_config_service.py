from typing import Callable
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import List, Any, Tuple, Dict
import json, os, datetime
from types import SimpleNamespace
from datetime import date
from openpyxl.cell.cell import Cell

import common.constants as constants
import common.utils as utils
from common.errors import *

from domain.state import ConfigStateProxy
from domain.analysis_run_config import AnalysisRunConfig
from domain.cycle_config import CycleConfig
from domain.batch_config import BatchConfig
from domain.general_config import GeneralConfig
from domain.workflow_config import WorkflowConfig
from domain.script_parameter_config import ScriptParameterConfig

from repositories.link_type_repository import LinkTypeRepository

from services.object_registration_service import ObjectRegistrationService
from services.identifier_service import IdentifierService
from services.solution_service import SolutionService
from services.request_service import RequestService


class BatchConfigService:
    
    # private members  
    def __init__(self, 
                 state : ConfigStateProxy,
                 object_registration_service : ObjectRegistrationService,
                 solution_service : SolutionService,
                 link_type_repository : LinkTypeRepository,
                 identifier_service : IdentifierService,
                 request_service : RequestService
                ) -> None:
        if state is None: raise ValueError(f"state cannot be empty")
        if object_registration_service is None: raise ValueError(f"object_registration_service cannot be empty")
        if link_type_repository is None: raise ValueError(f"link_type_repository cannot be empty")
        if identifier_service is None: raise ValueError(f"identifier_service cannot be empty")
        if solution_service is None: raise ValueError(f"solution_service cannot be empty")
        if request_service is None: raise ValueError(f"request_service cannot be empty")

        self._state = state
        self._object_registration_service = object_registration_service
        self._link_type_repository = link_type_repository
        self._identifier_service = identifier_service
        self._solution_service = solution_service
        self._request_service = request_service
    

    def __convert(self, value: Any, to_type : Any):
        if (to_type is None or len(str(to_type)) == 0): raise ValueError(f"to_type cannot be empty")

        if (to_type in (object, SimpleNamespace)):
            if (value is not None and len(str(value)) > 0):
                return json.loads(str(value), object_hook = lambda d: SimpleNamespace(**d))
            else: 
                return None
        elif (to_type == date):
            if isinstance(value, datetime.datetime):
                return str(value.date())
            elif isinstance(value, datetime.date):
                return str(value)
        elif (to_type == datetime.datetime):
            if isinstance(value, datetime.datetime):
                return value.isoformat()
            elif isinstance(value, datetime.date):
                return datetime.datetime.combine(value, datetime.datetime.min.time()).isoformat()
            
        return value


    # protected members
    def _parse_config_by_key_value(self, 
                                   sheet: Worksheet,
                                   config_create_func: Any
                                   ) -> Any:

        if sheet is None: raise ValueError(f"sheet cannot be empty")

        config : GeneralConfig = config_create_func()
        attributes = utils.get_attributes(config)
        column_map = {}

        for col in sheet.iter_cols(max_row = 1):
            headerCell = col[0]

            if (headerCell.value in ["parameter", "value"]):
                col_idx = headerCell.col_idx - 1
                column_map.update({ headerCell.value: col_idx })
        
        for row in sheet.iter_rows(min_row = 2):

            parameter_name = row[column_map["parameter"]].value
            if parameter_name in attributes:

                parameter_value = row[column_map["value"]].value
                value = self.__convert(
                    value = parameter_value, 
                    to_type = attributes[parameter_name])
                setattr(config, parameter_name, value)

        return config
    

    def _create_config_runnable(self, 
                                sheet_name : str,
                                excel_row : Tuple[Cell, ...], 
                                column_fields_map : Dict[str, int], 
                                field_type_map : Dict[str, str],
                                column_other_map : Dict[str, int],
                                column_links_map : Dict[str, int]
                               ) -> Any:
        
        if sheet_name == constants.SHEET_NAME_CYCLES:
            cycle_config = CycleConfig(ordinal_number = excel_row[0].row)
            cycle_config.validate_headers(column_fields_map, column_other_map)
            cycle_config.set_from_excel_row(
                excel_row, 
                column_fields_map,
                field_type_map,
                column_other_map,
                column_links_map)

            return cycle_config

        if sheet_name == constants.SHEET_NAME_ANALYSIS_RUNS:
            ar_config = AnalysisRunConfig(ordinal_number = excel_row[0].row)
            ar_config.validate_headers(column_fields_map, column_other_map)
            ar_config.set_from_excel_row(
                excel_row, 
                column_fields_map,
                field_type_map,
                column_other_map,
                column_links_map)

            return ar_config
        
        raise ValueError(f"Configuration for sheet_name '{sheet_name}' is not supported")
    

    def _parse_config_runnable(self, 
                               sheet : Worksheet
                              ) -> List[Any]:
        result = []
        column_comment_idx = -1
        column_fields_map : Dict[str, int]= {}
        column_other_map : Dict[str, int] = {}
        column_links_map : Dict[str, int] = {}
        rest_path = constants.SHEET_NAME_REST_PATH_MAP[sheet.title]

        field_definitions = self._object_registration_service.get_field_definitions(rest_path)
        field_type_map = dict([(field_def.name, field_def.type) for field_def in field_definitions])
        field_type_map.update(constants.CIRRUS_OBJECT_ROOT_PROPERTIES)

        object_registration = self._object_registration_service.get_object_registration(rest_path)
        # NOTE: is not cached
        link_types = self._link_type_repository.get_all_by_object_type_key(object_registration.key)
        link_type_identifier_keys = set([self._identifier_service.create_by_id_ssc(link_type.objectId, link_type.sourceSystemCd).get_key() \
                                         for link_type in link_types])

        for col in sheet.iter_cols(max_row = 1):
            headerCell = col[0]

            link_identifier = self._identifier_service.create_by_key(str(headerCell.value))

            if (headerCell.value in field_type_map):
                column_fields_map.update({ str(headerCell.value): headerCell.col_idx - 1 })
            elif (headerCell.value == constants.SHEET_COLUMN_COMMENT):
                column_comment_idx = headerCell.col_idx - 1
            elif (link_identifier.get_key() in link_type_identifier_keys):
                column_links_map.update({ link_identifier.get_key(): headerCell.col_idx - 1 })
            else:
                column_other_map.update({ str(headerCell.value): headerCell.col_idx - 1 })

        for row in sheet.iter_rows(min_row = 2):
            
            # skip commented out row
            if (column_comment_idx >= 0 and str(row[column_comment_idx].value) == constants.SHEET_COLUMN_COMMENT):
                continue

            config = self._create_config_runnable(
                sheet_name = sheet.title,
                excel_row = row, 
                column_fields_map = column_fields_map, 
                field_type_map = field_type_map, 
                column_other_map = column_other_map,
                column_links_map = column_links_map)
            
            result.append(config)
        
        return result
    

    def _parse_config_parameter(self, 
                                sheet : Worksheet,
                                config_create_func: Callable[[], ScriptParameterConfig | WorkflowConfig]
                               ) -> List[Any]:
        result = []
        column_fields_map : Dict[str, int]= {}
        column_other_map : Dict[str, int] = {}
        column_comment_idx = -1
        instance = config_create_func()
        attributes = utils.get_attributes(instance)
        
        for col in sheet.iter_cols(max_row = 1):
            headerCell = col[0]

            if (headerCell.value in attributes.keys()):
                column_fields_map.update({ str(headerCell.value): headerCell.col_idx - 1 })
            elif (headerCell.value == constants.SHEET_COLUMN_COMMENT):
                column_comment_idx = headerCell.col_idx - 1
            else:
                column_other_map.update({ str(headerCell.value): headerCell.col_idx - 1 })

        instance.validate_headers(
            sheet_name = sheet.title,
            column_fields_map = column_fields_map, 
            column_other_map = column_other_map)

        for row in sheet.iter_rows(min_row = 2):
            
            # skip commmented out row
            if (column_comment_idx >= 0 and str(row[column_comment_idx].value) == constants.SHEET_COLUMN_COMMENT):
                continue

            config = config_create_func()
            config.set_from_excel_row(
                excel_row = row,
                column_fields_map = column_fields_map)
            
            result.append(config)
        
        return result
    
    # def _validate_general_parameters(self, general_config : GeneralConfig):

    #     access_expiration_in_sec = self._request_service.get_access_expiration_from_issuedAt()
    #     if access_expiration_in_sec < general_config.workflow_wait_sleep:
    #         raise ValueError(f"workflow_wait_sleep general configuration value "
    #                          f"{general_config.workflow_wait_sleep} sec cannot exceed access token expiration duration "
    #                          f"({access_expiration_in_sec} sec).")
    #     if access_expiration_in_sec < general_config.script_wait_sleep:
    #         raise ValueError(f"script_wait_sleep general configuration value " 
    #                          f"{general_config.script_wait_sleep} sec cannot exceed access token expiration duration " 
    #                          f"({access_expiration_in_sec} sec).")
    
    
    def _validate_duplicate_script_parameters(self, 
                                              configs : List[ScriptParameterConfig],
                                              sheet_name : str
                                             ):
        key_count = {}
        duplicates = {}
        for c in configs:
           key = c.get_key_unique()
           key_count[key] = key_count.get(key, 0) + 1
           if key_count[key] > 1:
               duplicates[key] = f"{c.objectId or ''}:{c.sourceSystemCd or ''}:{c.task_name or ''}:{c.parameter_name or ''}:{c.parent_parameter or ''}"

        if len(duplicates) > 0:
            new_line = "\n"
            raise ParseError(f"The following duplicated cycle script parameters were found ('{sheet_name}' worksheet): " \
                             f"\n{new_line.join(duplicates.values())}.")
        

    def _validate_cyclic_references(self,
                                    iterations : set,
                                    parameter_to_validate : ScriptParameterConfig | None,
                                    parameters_by_group : List[ScriptParameterConfig],
                                    sheet_name : str
                                    ):

        if parameter_to_validate is None:           
            children = [c for c in parameters_by_group if c.parent_parameter is not None]
            for child in children:
                self._validate_cyclic_references(set(), child, parameters_by_group, sheet_name)
        else:
            iteration_step = f"{parameter_to_validate.parameter_name}:{parameter_to_validate.parent_parameter}"
            if (iteration_step in iterations):
                raise ValueError(f"Cyclic script parameter reference detected ('{sheet_name}' worksheet): " \
                             f"\n{iteration_step}")
            
            iterations.add(iteration_step)
            if parameter_to_validate.parent_parameter is not None \
                or len(str(parameter_to_validate.parent_parameter)) > 0:
                
                parents = [c for c in parameters_by_group \
                        if c.parameter_name == parameter_to_validate.parent_parameter]
                for parent in parents:
                    self._validate_cyclic_references(iterations, parent, parameters_by_group, sheet_name)


    def _validate_workflow_parameter_sets(self,
                                          workflow_configs : List[WorkflowConfig],
                                          cycle_script_parameter_configs : Dict[str, Dict[str, List[ScriptParameterConfig]]],
                                          sheet_name : str
                                         ) -> None:
        new_line = "\n"
        unexisting_param_sets : List[str] = []
        run_script_transition_name = self._solution_service.get_cycle_run_script_transition_name()

        for config in workflow_configs:
            if str(config.transition_name).upper() == str(run_script_transition_name).upper():
                task_name_param_set_key = f"{config.task_name}:{config.parameter_set or ''}"
                config_key = config.get_key()
                if config_key in cycle_script_parameter_configs:
                    parameter_configs = cycle_script_parameter_configs[config_key]
                    if not task_name_param_set_key in parameter_configs:
                        unexisting_param_sets.append(f"{config.task_name}:{config.parameter_set or ''}")
        
        if len(unexisting_param_sets) > 0:
            raise ValueError(
                f"Workflow configuration references the following unexisting parameter sets ('{sheet_name}' worksheet): " \
                f"\ntask_name:parameter_set" \
                f"\n{new_line.join(unexisting_param_sets)}")


    def _validate_script_parameters_references(self,
                                               configs : List[ScriptParameterConfig],
                                               sheet_name : str):
        groups = set([c.get_key_group() for c in configs])
        new_line = "\n"
        for group in groups:
            parameters_by_group = [c for c in configs if c.get_key_group() == group]
            unexisting_parents = [f"{p.parameter_name} : {p.parent_parameter}" for p in parameters_by_group \
                                  if p.parent_parameter is not None and p.parent_parameter not in \
                                    [p2.parameter_name for p2 in parameters_by_group]]
            if len(unexisting_parents) > 0:
                raise ValueError(f"The following script parameters reference unexisting parent parameters ('{sheet_name}' worksheet): " \
                                f"\n{new_line.join(unexisting_parents)}")
            self._validate_cyclic_references(set(), None, parameters_by_group, sheet_name)


    def _map_script_parameters_by_config_key_task_name(self, 
                                                       configs : List[ScriptParameterConfig]
                                                      ) -> Dict[str, Dict[str, List[ScriptParameterConfig]]]:
        result : Dict[str, Dict[str, List[ScriptParameterConfig]]] = {}
        config_keys = set([config.get_key() for config in configs])

        for key in config_keys:
            task_name_param_sets = set([(config.task_name, config.parameter_set) for config in configs if config.get_key() == key])
            configs_by_task_name : Dict[str, List[ScriptParameterConfig]] = {}

            for task_name_param in task_name_param_sets:
                task_name = task_name_param[0]
                param_set = task_name_param[1]
                task_name_param_set_key = f"{task_name}:{param_set or ''}"
                config_list = [config for config in configs \
                               if config.task_name == task_name \
                                and config.parameter_set == param_set \
                                    and config.get_key() == key]
                configs_by_task_name.update({ task_name_param_set_key: config_list })
            
            result.update({ key : configs_by_task_name })

        return result
    

    def _map_script_parameters_by_config_key_dist_task_name(self,
                                                            configs : List[ScriptParameterConfig]
                                                           ) -> Dict[str, Dict[str, ScriptParameterConfig]]:
        result : Dict[str, Dict[str, ScriptParameterConfig]] = {}
        config_keys = set([config.get_key() for config in configs])

        for key in config_keys:
            task_names = set([config.task_name for config in configs if config.get_key() == key])
            configs_by_task_name : Dict[str, ScriptParameterConfig] = {}

            for task_name in task_names:

                config_list = [config for config in configs if config.task_name == task_name and config.get_key() == key]
                if (len(config_list) > 1): 
                    raise ParseError(f"Duplicated task name '{task_name}' found in script parameters.")
                elif (len(config_list) == 1):
                    configs_by_task_name.update({ task_name: config_list[0] })
                
            result.update({ key : configs_by_task_name })

        return result


    def _map_script_parameters_by_config_key(self,
                                             configs : List[ScriptParameterConfig]
                                            ) -> Dict[str, List[ScriptParameterConfig]]:
        result : Dict[str, List[ScriptParameterConfig]] = {}
        config_keys = set([config.get_key() for config in configs])

        for key in config_keys:
            config_list = [config for config in configs if config.get_key() == key]
            result.update({ key : config_list })

        return result
    
    def _validate_duplicate_workflow_parameter(self, 
                                               configs : List[WorkflowConfig],
                                               sheet_name : str
                                              ):
        key_count = {}
        duplicates = {}
        for c in configs:
           key = c.get_key_unique()
           key_count[key] = key_count.get(key, 0) + 1
           if key_count[key] > 1:
               duplicates[key] = f"{c.objectId or ''}:{c.sourceSystemCd or ''}:{c.task_name or ''}:{c.iteration or ''}"

        if len(duplicates) > 0:
            new_line = "\n"
            raise ParseError(f"The following duplicated cycle workflow parameters were found ('{sheet_name}' worksheet): " \
                             f"\nobjectId:sourceSystemCd:task_name:iteration" \
                             f"\n{new_line.join(duplicates.values())}.")


    def _map_cycle_workflow_by_config_key_task_name(self,
                                                    configs : List[WorkflowConfig]
                                                   ) -> Dict[str, Dict[str, List[WorkflowConfig]]]:
        result : Dict[str, Dict[str, List[WorkflowConfig]]] = {}
        config_keys = set([config.get_key() for config in configs])

        for key in config_keys:
            task_names = set([config.task_name for config in configs if config.get_key() == key])
            configs_by_task_name : Dict[str, List[WorkflowConfig]] = {}

            for task_name in task_names:
                config_list = [config for config in configs \
                               if config.task_name == task_name \
                                and config.get_key() == key]
                config_list.sort(key = lambda i: str(i.iteration or '0')) # sort by iteration
                configs_by_task_name.update({ str(task_name): config_list })
                
            result.update({ key : configs_by_task_name })

        return result


    # public members
    def parse(self, 
              config_file_path : str
             ) -> BatchConfig:
        if config_file_path is None or len(str(config_file_path)) == 0: raise ValueError("config_file_path cannot be empty")
        if not os.path.exists(config_file_path): raise ValueError(f"Configuration file '{config_file_path}' does not exist")

        batch_workbook = load_workbook(filename = config_file_path)
        general_config : GeneralConfig = None
        analysis_run_configs : List[AnalysisRunConfig] = []
        analysis_run_script_parameter_configs : Dict[str, List[ScriptParameterConfig]] = {}
        cycle_configs : List[CycleConfig]= []
        cycle_script_parameter_configs_list : List[ScriptParameterConfig] = []
        cycle_script_parameter_configs : Dict[str, Dict[str, List[ScriptParameterConfig]]] = {}
        cycle_workflow_configs : Dict[str, Dict[str, List[WorkflowConfig]]] = {}
        sheets = [sheetname.lower() for sheetname in batch_workbook.sheetnames]

        if constants.SHEET_NAME_GENERAL in sheets:
            general_sheet = batch_workbook[constants.SHEET_NAME_GENERAL]
            general_config = self._parse_config_by_key_value(
                sheet = general_sheet, 
                config_create_func = lambda : GeneralConfig())
            # self._validate_general_parameters(general_config)

        if constants.SHEET_NAME_ANALYSIS_RUNS in sheets:
            analysis_run_sheet = batch_workbook[constants.SHEET_NAME_ANALYSIS_RUNS]
            analysis_run_configs = self._parse_config_runnable(
                sheet = analysis_run_sheet)
            analysis_run_configs = [config for config in analysis_run_configs \
                                    if not (not config.objectId or not config.sourceSystemCd)]
            
        if constants.SHEET_NAME_CYCLES in sheets:

            cycles_sheet = batch_workbook[constants.SHEET_NAME_CYCLES]
            cycle_configs : List[CycleConfig] = self._parse_config_runnable(
                sheet = cycles_sheet)
            cycle_configs = [config for config in cycle_configs \
                             if not (not config.objectId or not config.sourceSystemCd)]
        
        if constants.SHEET_NAME_CYCLE_SCRIPT_PARAMETERS in sheets:
            cycle_script_parameters_sheet = batch_workbook[constants.SHEET_NAME_CYCLE_SCRIPT_PARAMETERS]
            rest_path = constants.SHEET_NAME_REST_PATH_MAP[cycle_script_parameters_sheet.title]
            script_params_configs : List[ScriptParameterConfig] = self._parse_config_parameter(
                sheet = cycle_script_parameters_sheet, 
                config_create_func = lambda : ScriptParameterConfig(rest_path, constants.OBJECT_TYPE_CYCLE))
            cycle_script_parameter_configs_list = [config for config in script_params_configs \
                        if not (not config.objectId or not config.sourceSystemCd or not config.task_name)]
            self._validate_duplicate_script_parameters(
                cycle_script_parameter_configs_list,
                constants.SHEET_NAME_CYCLE_SCRIPT_PARAMETERS)
            self._validate_script_parameters_references(
                cycle_script_parameter_configs_list,
                constants.SHEET_NAME_CYCLE_SCRIPT_PARAMETERS)
            cycle_script_parameter_configs = self._map_script_parameters_by_config_key_task_name(script_params_configs)
        
        if constants.SHEET_NAME_ANALYSIS_RUN_SCRIPT_PARAMETERS in sheets:
            analysis_run_script_parameters_sheet = batch_workbook[constants.SHEET_NAME_ANALYSIS_RUN_SCRIPT_PARAMETERS]
            rest_path = constants.SHEET_NAME_REST_PATH_MAP[analysis_run_script_parameters_sheet.title]
            script_params_configs : List[ScriptParameterConfig] = self._parse_config_parameter(
                sheet = analysis_run_script_parameters_sheet, 
                config_create_func = lambda : ScriptParameterConfig(rest_path, constants.OBJECT_TYPE_ANALYSIS_RUN))
            script_params_configs = [config for config in script_params_configs \
                        if not (not config.objectId or not config.sourceSystemCd)]
            self._validate_duplicate_script_parameters(
                script_params_configs,
                constants.SHEET_NAME_ANALYSIS_RUN_SCRIPT_PARAMETERS)
            self._validate_script_parameters_references(
                script_params_configs,
                constants.SHEET_NAME_ANALYSIS_RUN_SCRIPT_PARAMETERS)
            analysis_run_script_parameter_configs = self._map_script_parameters_by_config_key(configs = script_params_configs)
        
        # workflow should go after script parameters as there is reference to the parameter_set
        if constants.SHEET_NAME_CYCLE_WORKFLOW in sheets:
            cycle_workflow_sheet = batch_workbook[constants.SHEET_NAME_CYCLE_WORKFLOW]
            rest_path = constants.SHEET_NAME_REST_PATH_MAP[cycle_workflow_sheet.title]
            workflow_configs : List[WorkflowConfig] = self._parse_config_parameter(
                sheet = cycle_workflow_sheet, 
                config_create_func = lambda : WorkflowConfig(rest_path, constants.OBJECT_TYPE_CYCLE))
            workflow_configs = [config for config in workflow_configs \
                        if not (not config.objectId or not config.sourceSystemCd or not config.task_name)]
            self._validate_duplicate_workflow_parameter(
                workflow_configs,
                constants.SHEET_NAME_CYCLE_WORKFLOW
                )
            self._validate_workflow_parameter_sets(
                workflow_configs,
                cycle_script_parameter_configs,
                constants.SHEET_NAME_CYCLE_WORKFLOW
                )
            cycle_workflow_configs = self._map_cycle_workflow_by_config_key_task_name(workflow_configs)
        

        return BatchConfig(file_path = config_file_path,
                           general_config = general_config, 
                           cycle_configs = cycle_configs, 
                           cycle_script_parameter_configs = cycle_script_parameter_configs,
                           cycle_workflow_configs = cycle_workflow_configs,
                           analysis_run_configs = analysis_run_configs,
                           analysis_run_script_parameter_configs = analysis_run_script_parameter_configs
                           )
    

    def get_from_state(self) -> BatchConfig:
        return self._state.get()
    

    def put_config_to_state(self, config : BatchConfig) -> None:
        self._state.lock()

        try:
            self._state.update(config)
        finally:
            self._state.unlock()
